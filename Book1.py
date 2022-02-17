# ステップ1｜ライブラリの設定


import os
from datetime import datetime, date

import files as files
import openpyxl
import openpyxl as excel
import openpyxl as px
import pandas as pd

from openpyxl import load_workbook

wb = "Book1.xlsx"
wb = excel.Workbook()
wb.save(r'C:\Users\ists\PycharmProjects\Book1\Book1.xlsx')

path = (r'C:\Users\ists\PycharmProjects\Book1\Book1.xlsx')
print(files)
ss = load_workbook(path)
# sheet1 = ss['sheet1']
# カレントディレクトリのフォルダやファイルを全て取得
path = os.getcwd()
files = os.listdir(path)
print(files)

# フォルダやファイルを全て取得

df = pd.read_excel('Book1.xlsx', engine='openpyxl')

df = pd.read_excel(r'C:\Users\ists\PycharmProjects\Book1\Book1.xlsx', sheet_name="集計")
print(df)
df = pd.read_excel('Book1.xlsx', sheet_name=['データ', '集計'])

# ステップ2｜所定フォルダ内の「Book1.xlsx」を指定して読み込む
os.path.exists('Book1.xlsx')
os.path.exists('Python')
path = r'C:\Users\ists\PycharmProjects\Book1\Book1.xlsx'
wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(r'C:\Users\ists\PycharmProjects\Book1\Book1.xlsx')

ws = wb['データ']
os.path.basename("Book1.xlsx")
sheet = wb.get_sheet_by_name('データ')
sheet = wb.active

# ワークシートの作成
# wb.create_sheet(title='データ')
# wb.create_sheet(title='集計')
wb.create_sheet(index=0, title='データ')

#シート名を変更
ws.title = 'データ'

#別名で保存
wb.save(r'C:\Users\ists\PycharmProjects\データ\Book1.xlsx')
wb1 = px.load_workbook('データ')
wb2 = px.load_workbook(r"C:\Users\ists\PycharmProjects\Book1\Book1.xlsx")
ws1 = wb1['データ']
ws2 = wb2['集計']
ws = wb["データ"]
ws = wb["集計"]
print(ws)

# ステップ3｜集計範囲の取得
# 今日
dt = datetime
today = datetime.date
datetime(2022, 2, 1, 10, 0)
startDate = datetime.date
datetime(2022, 2, 1, 10, 0)
datetime.today()
print(datetime(2022, 2, 1, 9, 55, 28))
print(dt)
print(today)

# # 日付と時刻を構成する要素の取り出し
# print(f'year: {dt.year}, month: {dt.month}, day: {dt.day}')
# print(f'hour: {dt.hour}, minute: {dt.minute}, second: {dt.second}')
# print(f'micro second: {dt.microsecond}')
#
# # datetimeオブジェクトから日付または時刻を取り出す
# d = dt.date()
# print(d)
#
# t = dt.time()
# print(t)

# スタート日、時間を設定
t = datetime.now()
datetime = datetime.date

now = datetime

print('datetime.datetime.now()')
print(f'{t}\n')

print('# f-string')
print(f'year: {t.year}')
print(f'month: {t.month}')
print(f'day: {t.day}')
print(f'hour: {t.hour}')
print(f'minute: {t.minute}')
print(f'second: {t.second}')
print(f'microsecond: {t.microsecond}')
print(f'tzinfo: {t.tzinfo}\n')

print('# str.format()')
print('year: {t.year}'.format(t=t))
print('month: {t.month}'.format(t=t))
print('day: {t.day}'.format(t=t))
print('hour: {t.hour}'.format(t=t))
print('minute: {t.minute}'.format(t=t))
print('second: {t.second}'.format(t=t))
print('microsecond: {t.microsecond}'.format(t=t))
print('tzinfo: {t.tzinfo}\n'.format(t=t))

print('# % 演算子')
print('year: %d' % t.year)
print('month: %d' % t.month)
print('day: %d' % t.day)
print('hour: %d' % t.hour)
print('minute: %d' % t.minute)
print('second: %d' % t.second)
print('microsecond: %d' % t.microsecond)
print('tzinfo: %s\n' % t.tzinfo)

start = date()
end = datetime('2022-02-10').date()
today = datetime.date.today()
print(today.strftime('%Y%m%d'))

# 日数差を表示
print((end - start).days)


def daterange(_start, _end):
    for n in range((_end - _start).days):
        day = (_start + datetime.timedelta(n)).strftime('%y%y/%m/%d')
        yield day


period = []
for i in daterange(start, end):
    period.append(i)

# 差分を計算
period = ws1 - startDate
print(period)
startdate = datetime.datetime(int(ws2['B2'].value), int(ws2['C2'].value), int(ws2['D2'].value))
enddate = datetime.datetime(int(ws2['B3'].value), int(ws2['C3'].value), int(ws2['D3'].value))

# ステップ4｜エクセルの最終行や最終列の取得
lastrow1 = ws1.max_row
lastrow2 = ws2.max_row
lastcol2 = ws2.max_column

# ステップ5｜「データ」シートを読み込み、2次元配列values1として取得
values1 = [[cell.value for cell in row1] for row1 in ws1]

# ステップ6｜条件に合う情報をFor文とIF文で場合分けして取得
for i in range(7, lastrow2 + 1):
    for j in range(2, lastcol2 + 1):
        counter = 0

        for k in range(1, lastrow1):
            if values1[k][1] == ws2.cell(row=i, column=1).value:
                if values1[k][2] == ws2.cell(row=6, column=j).value:
                    torihikidate = values1[k][3]
                    if startdate <= torihikidate <= enddate:
                        kingaku = values1[k][4]
                        counter = counter + int(kingaku)

# ステップ7｜ステップ6に合わない場合に0を入れる
if counter is None:
    counter = 0

# ステップ8｜「集計」シートに結果を書き出す
ws2.cell(row=i, column=j).value = counter

# Program to read the entire file (absolute path) using read() function
file = open("python.txt", "r")
content = file.read()
print(content)
file.close()

# ステップ9｜「Book2.xlsx」として所定のフォルダに保存する
newfilepath = r"C:/Users/ists/OneDrive/デスクトップ/python/Book2.xlsx"
wb.save(newfilepath)
